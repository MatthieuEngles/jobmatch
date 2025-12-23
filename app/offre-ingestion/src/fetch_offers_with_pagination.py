import json
import os
import re
import sys
import time
from datetime import datetime, UTC, date, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook


# ----------------------------
# Utilitaires .env (sans dépendance externe)
# ----------------------------
def load_dotenv(dotenv_path: Path) -> None:
    """
    Charge un .env simple (KEY=VALUE) dans os.environ si la variable n'existe pas déjà.
    - Ignore lignes vides et commentaires (#)
    - Supporte valeurs entre guillemets simples/doubles
    """
    if not dotenv_path.exists():
        return

    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def require_env(name: str) -> str:
    val = os.environ.get(name, "").strip()
    if not val:
        print(f"Erreur: variable d'environnement manquante: {name}")
        sys.exit(1)
    return val


# ----------------------------
# Gestion de la date à requêter
# ----------------------------
def parse_target_date(argv: list[str]) -> date:
    """
    - Sans argument: prend la veille (UTC)
    - Avec argument: attend YYYY-MM-DD et prend ce jour-là
    """
    if len(argv) <= 1:
        return datetime.now(timezone.utc).date() - timedelta(days=1)

    s = argv[1]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        print(
            "Erreur: argument date invalide. Format attendu: YYYY-MM-DD (ex: 2025-12-20)"
        )
        sys.exit(1)


# ----------------------------
# Chargement config
# ----------------------------
PROJECT_ROOT = (
    Path(__file__).resolve().parents[1]
)  # .../scripts/fetch_offers.py -> racine projet
load_dotenv(PROJECT_ROOT / ".env")

# Obligatoires
CLIENT_ID = require_env("FT_CLIENT_ID")
CLIENT_SECRET = require_env("FT_CLIENT_SECRET")

# Optionnelles (avec défauts)
SCOPE = os.environ.get("FT_SCOPE", "api_offresdemploiv2 o2dsoffre").strip()
OAUTH_URL = os.environ.get(
    "FT_OAUTH_URL",
    "https://entreprise.francetravail.fr/connexion/oauth2/access_token?realm=/partenaire",
).strip()
API_URL_BASE = os.environ.get(
    "FT_API_URL_BASE",
    "https://api.francetravail.io/partenaire/offresdemploi/v2/offres/search",
).strip()
LOGS_XLSX = os.environ.get("FT_LOGS_XLSX", "logs.xlsx").strip()

DEFAULT_ROMECODES_PATH = PROJECT_ROOT / "src" / "data_persist" / "rome_codes.txt"
ROMECODES_PATH = Path(
    os.environ.get("FT_ROMECODES_PATH", str(DEFAULT_ROMECODES_PATH))
).resolve()

# Sécurité token
TOKEN_SKEW_SECONDS = 60

_token_cache = {
    "access_token": None,
    "expires_at": 0.0,  # epoch seconds
}


def _now_epoch() -> float:
    return time.time()


def get_token(force_refresh: bool = False) -> str:
    """
    Récupère un token valide (cache + refresh).
    """
    if (
        not force_refresh
        and _token_cache["access_token"]
        and _now_epoch() < (_token_cache["expires_at"] - TOKEN_SKEW_SECONDS)
    ):
        return _token_cache["access_token"]

    data = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": SCOPE,
    }

    r = requests.post(
        OAUTH_URL,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data=data,
        timeout=30,
    )
    r.raise_for_status()
    payload = r.json()

    access_token = payload["access_token"]
    expires_in = int(payload.get("expires_in", 0))

    _token_cache["access_token"] = access_token
    _token_cache["expires_at"] = _now_epoch() + expires_in

    return access_token


def auth_headers() -> dict:
    return {
        "Authorization": f"Bearer {get_token()}",
        "Accept": "application/json",
    }


def load_rome_codes(path: Path) -> list[str]:
    if not path.exists():
        print(f"Erreur: fichier ROMECODES introuvable: {path}")
        sys.exit(1)

    codes: list[str] = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        codes.append(line)
    if not codes:
        print(f"Erreur: aucun ROMECODE dans {path}")
        sys.exit(1)
    return codes


def extract_total_from_content_range(content_range: str | None) -> int | None:
    # Exemple: "offres 0-149/591250"
    if not content_range:
        return None
    m = re.search(r"/\s*(\d+)\s*$", content_range)
    return int(m.group(1)) if m else None


def init_logs_xlsx(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb["logs"] if "logs" in wb.sheetnames else wb.create_sheet("logs")
        if ws.max_row == 0:
            ws.append(
                [
                    "timestamp_utc",
                    "romeCode",
                    "range",
                    "http_status",
                    "offers_returned",
                    "total_header",
                ]
            )
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "logs"
    ws.append(
        [
            "timestamp_utc",
            "romeCode",
            "range",
            "http_status",
            "offers_returned",
            "total_header",
        ]
    )
    wb.save(path)
    return wb, ws


def log_row(
    ws, rome: str, range_str: str, status: int, offers_n: int, total: int | None
):
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    ws.append(
        [ts, rome, range_str, status, offers_n, total if total is not None else ""]
    )


def get_with_auto_refresh(
    session: requests.Session, url: str, params: dict, timeout: int = 30
) -> requests.Response:
    """
    Fait un GET.
    Si 401 -> refresh token -> retente 1 fois.
    """
    session.headers.update(auth_headers())
    r = session.get(url, params=params, timeout=timeout)

    if r.status_code != 401:
        return r

    get_token(force_refresh=True)
    session.headers.update(auth_headers())
    return session.get(url, params=params, timeout=timeout)


def write_offers_json_atomic(path: str, offers: list[dict]) -> None:
    """
    Ecrit 1 seule fois (rapide) et de manière atomique:
    - écrit dans un fichier temporaire
    - puis remplace le fichier final
    """
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump({"resultats": offers}, f, ensure_ascii=False)
    os.replace(tmp_path, path)


def main() -> int:
    debut = datetime.now(timezone.utc)

    target_date = parse_target_date(sys.argv)
    min_dt = f"{target_date.isoformat()}T00:00:00Z"
    max_dt = f"{(target_date + timedelta(days=1)).isoformat()}T00:00:00Z"

    api_url = f"{API_URL_BASE}?minCreationDate={min_dt}&maxCreationDate={max_dt}"
    out_json = f"offer_{target_date.isoformat()}.json"
    rome_codes = load_rome_codes(ROMECODES_PATH)

    print(f"ROMECODES: {len(rome_codes)} (depuis {ROMECODES_PATH})")
    print(f"Requête période: {min_dt} -> {max_dt}")
    print(f"URL: {api_url}")
    print(f"Fichier de sortie: {out_json}")
    print(f"Logs: {LOGS_XLSX}")

    wb_logs, ws_logs = init_logs_xlsx(LOGS_XLSX)
    session = requests.Session()

    # IMPORTANT: on accumule toutes les offres ici (une seule écriture fichier à la fin)
    all_offers: list[dict] = []

    # Throttle: assurer au moins 0.11s entre deux requêtes API (début->début)
    MIN_API_INTERVAL = 0.11
    last_api_call_ts: float | None = None  # time.monotonic()

    for i, rome in enumerate(rome_codes, start=1):
        start = 0
        step = 150

        while True:
            range_str = f"{start}-{start + step - 1}"
            params = {"codeROME": rome, "range": range_str}

            try:
                # Assure l'intervalle minimum entre deux appels API (avant d'appeler l'API)
                now_ts = time.monotonic()
                if last_api_call_ts is not None:
                    elapsed = now_ts - last_api_call_ts
                    remaining = MIN_API_INTERVAL - elapsed
                    if remaining > 0:
                        time.sleep(remaining)

                r = get_with_auto_refresh(session, api_url, params=params, timeout=30)
                # Marque le début de CET appel API
                last_api_call_ts = time.monotonic()
                status = r.status_code
                total = extract_total_from_content_range(r.headers.get("Content-Range"))

                if status == 204:
                    log_row(ws_logs, rome, range_str, status, 0, total)
                    print(
                        f"{i}/{len(rome_codes)} {rome} range={range_str} -> 204 (0 offre)"
                    )
                    break

                if status in (200, 206):
                    payload = r.json() if r.content else {}
                    offers = payload.get("resultats") or []
                    offers_n = len(offers)

                    # Au lieu d'écrire dans le fichier à chaque boucle: on stocke en mémoire
                    if offers:
                        all_offers.extend(offers)

                    log_row(ws_logs, rome, range_str, status, offers_n, total)
                    print(
                        f"{i}/{len(rome_codes)} {rome} range={range_str} -> {status} ({offers_n} offres) total={total}"
                    )

                    if status == 206:
                        start += step
                        continue

                    break

                log_row(ws_logs, rome, range_str, status, 0, total)
                print(
                    f"{i}/{len(rome_codes)} {rome} range={range_str} -> HTTP {status} (on skip)"
                )
                break

            except Exception as e:
                log_row(ws_logs, rome, range_str, 0, 0, None)
                print(
                    f"{i}/{len(rome_codes)} {rome} range={range_str} -> ERROR {e} (on skip)"
                )
                break

        wb_logs.save(LOGS_XLSX)

    wb_logs.save(LOGS_XLSX)

    # Une seule écriture JSON à la fin (beaucoup plus rapide)
    write_offers_json_atomic(out_json, all_offers)

    print(f"JSON écrit: {out_json} (offres: {len(all_offers)})")
    print(f"Logs écrits: {LOGS_XLSX}")
    fin = datetime.now(timezone.utc)
    duree = (fin - debut).total_seconds()
    print(f"Durée totale du script: {duree:.2f} secondes")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
